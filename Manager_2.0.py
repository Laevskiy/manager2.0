import tkinter as tk
import tkinter.ttk as ttk
import tkinter.messagebox as mb
import openpyxl
import datetime


win = tk.Tk()

win.title ("Photo_manager_2022")
win.geometry('520x550')
win.config(bg="#78DBE2")
win.resizable(False,False)

# 1 ФУНКЦИИ
def on_leave(btn):
    btn.bind ("<Leave>", func=lambda e: btn.config(
        fg= "Black"))
    btn.bind('<Enter>',func=lambda e: btn.config(
        fg= "White"))

def ochistka_01():
    city_old_1.grid_forget()
    name_old_1.grid_forget()
    tel_old_1.grid_forget()
    price_old_1.grid_forget()
    time_old_1.grid_forget()
    start_old_1.grid_forget()

def ochistka_02():
    city_lab.grid_forget()
    name_lab.grid_forget()
    tel_lab.grid_forget()
    price_lab.grid_forget()
    time_lab.grid_forget()
    start_lab.grid_forget()
    hint_3.grid_forget()
    btn_2_yes.grid_forget()
    btn_2_NO.grid_forget()
    btn_2_del.grid_forget()
    city_ent.grid_forget()
    name_ent.grid_forget()
    tel_ent.grid_forget()
    price_ent.grid_forget()
    time_ent.grid_forget()
    start_ent.grid_forget()
    btn_zapis.grid_forget()
    d_2.grid_forget()

def ochistka_03():
    hint_2.grid_forget()
    hint_3.grid_forget()
    btn_3_yes.grid_forget()
    btn_3_NO.grid_forget()

def zakritie():
    msg = "До скорых встреч"
    mb.showinfo("Информация", msg)
    win.destroy()
def data_del():
    btn_vvod["state"] = tk.NORMAL
    wb_list.delete_rows(b)
    wb.save(filename='D:\manager\photo_data_new.xlsx')
    msg = "Дата УДАЛЕНА!"
    mb.showinfo("Информация", msg)
    ochistka_01()
    ochistka_02()


def otrisovka_lab():
    city_lab.grid(row=4, column=0, padx=10, pady=10)
    name_lab.grid(row=5, column=0, padx=10, pady=10)
    tel_lab.grid(row=6, column=0, padx=10, pady=10)
    price_lab.grid(row=7, column=0, padx=10, pady=10)
    time_lab.grid(row=8, column=0, padx=10, pady=10)
    start_lab.grid(row=9, column=0, padx=10, pady=10)

def but_perezapis():
    hint_3.grid(row=10, column=0, padx=10, pady=10,columnspan=3)
    btn_2_yes.grid(row=11, column=0, padx=10, pady=10,stick = 'we')
    btn_2_NO.grid(row=11, column=1, padx=10, pady=10,stick = 'we')
    btn_2_del.grid(row=11, column=2, padx=10, pady=10,stick = 'we')

def otrisovka_suz_dan_per():
    data_del()
    ochistka_01()
    ochistka_02()
    otrisovka_dan_zap()

def otrisovka_suz_dan(city_old,name_old,tel_old,price_old,time_old,start_old):
    global city_old_1
    city_old_1 = tk.Label(win, text=city_old,
                          bg="#78DBE2",
                          font=("Arial", 10, "bold"),
                          )
    city_old_1.grid(row=4, column=1)

    global name_old_1
    name_old_1 = tk.Label(win, text = name_old,
                          bg="#78DBE2",
                          font=("Arial", 10, "bold"),
                          )
    name_old_1.grid(row=5, column=1)

    global tel_old_1
    tel_old_1 = tk.Label(win, text=tel_old,
                         bg="#78DBE2",
                         font=("Arial", 10, "bold"),
                         )
    tel_old_1.grid(row=6, column=1)

    global price_old_1
    price_old_1 = tk.Label(win, text=price_old,
                           bg="#78DBE2",
                           font=("Arial", 10, "bold"),
                           )
    price_old_1.grid(row=7, column=1)

    global time_old_1
    time_old_1 = tk.Label(win, text=time_old,
                          bg="#78DBE2",
                          font=("Arial", 10, "bold"),
                          )
    time_old_1.grid(row=8, column=1)

    global start_old_1
    start_old_1 = tk.Label(win, text=start_old,
                           bg="#78DBE2",
                           font=("Arial", 10, "bold"),
                           )
    start_old_1.grid(row=9, column=1)

def otrisovka_dan_zap():
    ochistka_03()
    city_lab.grid(row=6, column=0, padx=10, pady=10)
    name_lab.grid(row=7, column=0, padx=10, pady=10)
    tel_lab.grid(row=8, column=0, padx=10, pady=10)
    price_lab.grid(row=9, column=0, padx=10, pady=10)
    time_lab.grid(row=10, column=0, padx=10, pady=10)
    start_lab.grid(row=11, column=0, padx=10, pady=10)
    # Полей ввода
    city_ent.grid(row=6, column=1, padx=10, pady=10)
    name_ent.grid(row=7, column=1, padx=10, pady=10)
    tel_ent.grid(row=8, column=1, padx=10, pady=10)
    price_ent.grid(row=9, column=1, padx=10, pady=10)
    time_ent.grid(row=10, column=1, padx=10, pady=10)
    start_ent.grid(row=11, column=1, padx=10, pady=10)
    # Отрисовка кнопки записи
    btn_zapis.grid(row=12, column=0, columnspan=3, stick='we', padx=10, pady=10)
def zapis_new_data():
    btn_vvod["state"] = tk.NORMAL
    city_new = city_ent.get()
    name_new = name_ent.get()
    tel_new = tel_ent.get()
    price_new = price_ent.get()
    time_new = time_ent.get()
    start_new = start_ent.get()
    zap_data = (sercher, city_new,name_new,tel_new,price_new,time_new,start_new)
    wb_list.append(zap_data)
    wb.save('D:\manager\photo_data_new.xlsx')
    msg_2 = "Дата ЗАПИСАНА!"
    mb.showinfo("Информация", msg_2)
    ochistka_02()

def proverka():
    btn_vvod["state"] = tk.DISABLED
    day = combo_days.get()
    mouth = combo_mounth.get()
    year = combo_years.get()
    out_2 = f"{day} {mouth} {year}"
    out_2_1 = f"Дата: {out_2} уже ЗАБРОНИРОВАННА"
    out_2_2 = f"Дата: {out_2} СВОБОДНА! Хотите записать её?"
    global wb
    wb = openpyxl.reader.excel.load_workbook(filename='D:\manager\photo_data_new.xlsx')
    global wb_list
    wb_list = wb[year]
    # a - поиск количества строк
    a = wb_list.max_row
    mounth_num = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12)
    mouth_sl = dict(zip(mounth, mounth_num))
    # искомая датаsearcher
    global sercher
    sercher = datetime.datetime(int(year), mouth_sl[mouth], int(day))
    for i in range(2, a + 1):

        # print(wb_list['B'+str(i)].value, "дата:",wb_list['A'+str(i)].value )
        global val
        val = wb_list.cell(row=i, column=1).value

        if val == sercher:
            global b
            b = i
            print("Есть совпадение, по дате :", wb_list['A' + str(i)].value)
            city_old = wb_list['B'+str(i)].value
            name_old = wb_list['C'+str(i)].value
            tel_old  = wb_list['D'+str(i)].value
            price_old = wb_list['E'+str(i)].value
            time_old = wb_list['F'+str(i)].value
            start_old = wb_list['G'+str(i)].value

            # Отрисовка кнопок перезаписи
            global d_2
            d_2 = tk.Label(win, text= out_2_1,
                           font=("Arial", 10, "bold"),
                           bg = '#78DBE2',
                           fg="RED")
            d_2.grid(row=3, column=0, columnspan=3,padx=10, pady=10)
            otrisovka_suz_dan(city_old,name_old,tel_old,price_old,time_old,start_old)
            otrisovka_lab()
            but_perezapis()
            break

    else:

        d_2 = tk.Label(win, text=out_2_2,
                       font=("Arial", 10, "bold"),
                       bg='#78DBE2',
                       fg="RED")
        d_2.grid(row=3, column=0, columnspan=3,padx=10, pady=10)
        #hint_2.grid(row=3, column=0, padx=10, pady=10)
        btn_3_yes.grid(row = 5 , column=0, padx=10, pady=10,stick='we')
        btn_3_NO.grid(row = 5 , column=1, columnspan=2,padx=10, pady=10,stick='we')
#2 Кнопки/Лэйбы

# 2.1  Выпадающая дата
days = list(range(1, 32))
combo_days = ttk.Combobox(win,values=days, width = 4)
combo_days.current(0)
combo_days.grid(row = 1, column=0, padx=10, pady=10,stick = 'we')

#2.2  Выпадающий список с месяцами
mounth= ("ЯНВАРЬ","ФЕВРАЛЬ","МАРТ","АПРЕЛЬ","МАЙ","ИЮНЬ","ИЮЛЬ","АВГУСТ","СЕНТЯБРЬ","ОКТЯБРЬ","НОЯБРЬ","ДЕКАБРЬ")
combo_mounth = ttk.Combobox(win, values=mounth, width = 20)
combo_mounth.current(0)
combo_mounth.grid(row = 1, column=1, padx=10, pady=10,stick = 'w')

# 2.3 Выпадающий год
years = list(range(2022,2025))
combo_years = ttk.Combobox(win,values=years,  width = 10)
combo_years.current(0)
combo_years.grid(row = 1, column=2, padx=10, pady=10,stick = 'we')

# 2. 4 Кнопка забора даты / "Проверка даты"
btn_vvod = tk.Button(win,text= "Проверить дату",command= proverka,
                     padx=4, pady=4,
                     bg="#8affbd",
                     font=("Arial", 12,),
                     fg="BLACK",
                     )
btn_vvod.grid(row = 2, column=0,columnspan=3,stick = 'we',padx=10, pady=10)

# 2.5. Верхняя подсказка
hint_1 = tk.Label(win, text= "Выберите желаемую дату:",
                  bg="#78DBE2",
                  font=("Arial", 12, "bold"))
hint_1.grid(row = 0 , column=0, padx=10, pady=10)

# 2.6 Подсказак/ Записать дату ?
hint_2 = tk.Label(win, text= "Вы хотите записать дату?")

# 2. 7 Подсказка / Вы хотите перезаписать дату ?
hint_3 = tk.Label(win, text= "Вы хотите ПЕРЕЗАПИСАТЬ дату?",
                  bg="#78DBE2",
                  font=("Arial", 10, "bold"),
                  fg = 'Red'
                  )


#2.8 Кнопки выбора
btn_2_yes = tk.Button(win,text= "ДА",command= otrisovka_suz_dan_per,
                      font=("Arial", 10, "bold"),
                      bg="RED",
                      )

btn_2_NO = tk.Button(win,text= "Нет",command= zakritie,
                     font=("Arial", 10, "bold"),
                     bg="Green",
                     )
btn_2_del = tk.Button(win,text= "Удалить дату",command= data_del,
                      font=("Arial", 10, "bold"),
                      bg="RED",
                      )
#btn_3_zapis = tk.Button(win,text= "Записать дату",command= proverka,)
btn_3_yes = tk.Button(win,text= "ДА",command= otrisovka_dan_zap,
                      font=("Arial", 10, "bold"),
                      bg="GREEN",
                      )
btn_3_NO = tk.Button(win,text= "Нет",command= zakritie,
                     font=("Arial", 10, "bold"),
                     bg="RED",
                     )

# 2.9 Кнопки заполнения (город,начало)
# Лэйбы
city_lab = tk.Label(win, text= "Город:",
                    bg="#78DBE2",
                    font=("Arial", 10, "bold")
                    )
name_lab = tk.Label(win, text= "Имя заказчика:",
                    bg="#78DBE2",
                    font=("Arial", 10, "bold")
                    )
tel_lab = tk.Label(win, text= "Контактный телефон с кодом:",
                   bg="#78DBE2",
                   font=("Arial", 10, "bold")
                   )
price_lab = tk.Label(win, text= "Стоимость",
                     bg="#78DBE2",
                     font=("Arial", 10, "bold")
                     )
time_lab = tk.Label(win, text= "Продолжительность:",
                    bg="#78DBE2",
                    font=("Arial", 10, "bold")
                    )
start_lab = tk.Label(win, text= "Начало мероприятия:",
                     bg="#78DBE2",
                     font=("Arial", 10, "bold")
                     )
# Строки ввода данных
city_ent = tk.Entry(win)
name_ent = tk.Entry(win)
tel_ent = tk.Entry(win)
price_ent = tk.Entry(win)
time_ent = tk.Entry(win)
start_ent = tk.Entry(win)

# 2.10 Кнопка записи
btn_zapis = tk.Button(win,text= "Записать новую дату ",command= zapis_new_data,
                    padx=4, pady=4,
                    bg="#8affbd",
                    font=("Arial", 12,),
                    fg="BLACK")

on_leave(btn_vvod)
on_leave(btn_2_yes)
on_leave(btn_2_NO)
on_leave(btn_2_del)
on_leave(btn_3_yes)
on_leave(btn_3_NO)
win.mainloop()